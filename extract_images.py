"""Extract images from outdoor kitchen hot-selling PI Excel file"""
import openpyxl, os

wb = openpyxl.load_workbook(r"E:\CAB3060\02-工作\设计中\dd-deep-design\户外厨房\户外厨房热卖款PI(1).xlsx")
img_dir = r"E:\CAB3060\02-工作\设计中\dd-deep-design\assets\images"
os.makedirs(img_dir, exist_ok=True)

total = 0
for sn in wb.sheetnames:
    ws = wb[sn]
    safe_sn = sn.replace("-","").replace(" ","").replace("(","").replace(")","")
    if hasattr(ws, "_images") and ws._images:
        for idx, img in enumerate(ws._images):
            img_data = img._data()
            # Try to detect format
            if img_data[:4] == b'\x89PNG':
                ext = "png"
            elif img_data[:2] == b'\xff\xd8':
                ext = "jpg"
            else:
                ext = "png"
            fn = f"{safe_sn}-img-{idx+1}.{ext}"
            path = os.path.join(img_dir, fn)
            with open(path, "wb") as f:
                f.write(img_data)
            total += 1
            print(f"  {fn} ({len(img_data)} bytes)")

print(f"\nTotal images extracted: {total}")

# Also check the price file
wb2 = openpyxl.load_workbook(r"E:\CAB3060\02-工作\设计中\dd-deep-design\户外厨房\索而户外箱体价格-隋总(1).xlsx")
for sn in wb2.sheetnames:
    ws = wb2[sn]
    if hasattr(ws, "_images") and ws._images:
        print(f"\nPrice file - {sn}: {len(ws._images)} images")
        for idx, img in enumerate(ws._images):
            img_data = img._data()
            fn = f"suoer-price-img-{idx+1}.png"
            path = os.path.join(img_dir, fn)
            with open(path, "wb") as f:
                f.write(img_data)
            print(f"  {fn} ({len(img_data)} bytes)")
    else:
        print(f"\nPrice file - no embedded images")
